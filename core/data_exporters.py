from django.db.models import Q, F, Sum, Avg, Count
from django.utils import timezone
from django.conf import settings
from datetime import timedelta, datetime
import csv
import json
import logging
from io import StringIO, BytesIO
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows

from .models import Player, Match, PlayerStats, TeamStats, Formation, MatchLineup, MatchEvent, Analytics, Opponent

logger = logging.getLogger('core.exports')

class DataExporters:
    
    def __init__(self):
        self.logger = logging.getLogger('core.exports')
        self.export_path = settings.POWERBI_CONFIG['EXPORT_PATH']
        
    def export_for_powerbi(self, export_type='all', date_range=None):
        export_data = {
            'export_timestamp': timezone.now().isoformat(),
            'export_type': export_type,
            'data_sources': {}
        }
        
        if export_type in ['all', 'players']:
            export_data['data_sources']['players'] = self._export_players_data()
        
        if export_type in ['all', 'matches']:
            export_data['data_sources']['matches'] = self._export_matches_data(date_range)
        
        if export_type in ['all', 'performance']:
            export_data['data_sources']['player_performance'] = self._export_player_performance_data(date_range)
        
        if export_type in ['all', 'formations']:
            export_data['data_sources']['formations'] = self._export_formations_data(date_range)
        
        if export_type in ['all', 'analytics']:
            export_data['data_sources']['analytics'] = self._export_analytics_data(date_range)
        
        if export_type in ['all', 'team_stats']:
            export_data['data_sources']['team_statistics'] = self._export_team_stats_data(date_range)
        
        self._save_powerbi_export(export_data, export_type)
        
        self.logger.info(f"PowerBI export completed: {export_type}")
        return export_data
    
    def _export_players_data(self):
        players = Player.objects.filter(is_active=True)
        
        players_data = []
        for player in players:
            player_record = {
                'player_id': str(player.id),
                'squad_number': player.squad_number,
                'full_name': player.full_name,
                'first_name': player.first_name,
                'last_name': player.last_name,
                'position': player.position,
                'position_category': self._categorise_position(player.position),
                'preferred_foot': player.preferred_foot,
                'age': player.age,
                'height_cm': player.height,
                'weight_kg': player.weight,
                'market_value_gbp': float(player.market_value),
                'contract_expiry': player.contract_expiry.isoformat(),
                'fitness_level': player.fitness_level,
                'is_injured': player.is_injured,
                'created_date': player.created_at.isoformat(),
                'updated_date': player.updated_at.isoformat()
            }
            players_data.append(player_record)
        
        return {
            'table_name': 'players',
            'record_count': len(players_data),
            'columns': list(players_data[0].keys()) if players_data else [],
            'data': players_data
        }
    
    def _export_matches_data(self, date_range=None):
        matches_query = Match.objects.all()
        
        if date_range:
            start_date = datetime.fromisoformat(date_range['start_date'])
            end_date = datetime.fromisoformat(date_range['end_date'])
            matches_query = matches_query.filter(
                scheduled_datetime__date__range=[start_date.date(), end_date.date()]
            )
        
        matches = matches_query.select_related('opponent')
        
        matches_data = []
        for match in matches:
            match_record = {
                'match_id': str(match.id),
                'match_date': match.scheduled_datetime.date().isoformat(),
                'match_datetime': match.scheduled_datetime.isoformat(),
                'opponent_name': match.opponent.name,
                'opponent_league': match.opponent.league,
                'opponent_country': match.opponent.country,
                'match_type': match.match_type,
                'is_home_match': match.is_home,
                'venue': match.venue,
                'match_status': match.status,
                'chelsea_score': match.chelsea_score,
                'opponent_score': match.opponent_score,
                'goal_difference': match.chelsea_score - match.opponent_score,
                'result': match.result,
                'attendance': match.attendance or 0,
                'weather_conditions': match.weather_conditions,
                'referee': match.referee,
                'season': self._determine_season(match.scheduled_datetime),
                'month': match.scheduled_datetime.month,
                'day_of_week': match.scheduled_datetime.strftime('%A'),
                'kick_off_time': match.scheduled_datetime.time().isoformat()
            }
            matches_data.append(match_record)
        
        return {
            'table_name': 'matches',
            'record_count': len(matches_data),
            'columns': list(matches_data[0].keys()) if matches_data else [],
            'data': matches_data
        }
    
    def _export_player_performance_data(self, date_range=None):
        stats_query = PlayerStats.objects.select_related('player', 'match', 'match__opponent')
        
        if date_range:
            start_date = datetime.fromisoformat(date_range['start_date'])
            end_date = datetime.fromisoformat(date_range['end_date'])
            stats_query = stats_query.filter(
                match__scheduled_datetime__date__range=[start_date.date(), end_date.date()]
            )
        
        performance_data = []
        for stats in stats_query:
            performance_record = {
                'performance_id': f"{stats.player.id}_{stats.match.id}",
                'player_id': str(stats.player.id),
                'player_name': stats.player.full_name,
                'squad_number': stats.player.squad_number,
                'position': stats.player.position,
                'position_category': self._categorise_position(stats.player.position),
                'match_id': str(stats.match.id),
                'match_date': stats.match.scheduled_datetime.date().isoformat(),
                'opponent': stats.match.opponent.name,
                'is_home': stats.match.is_home,
                'match_result': stats.match.result,
                'minutes_played': stats.minutes_played,
                'player_rating': float(stats.rating),
                'goals': stats.goals,
                'assists': stats.assists,
                'total_goal_involvements': stats.goals + stats.assists,
                'passes_completed': stats.passes_completed,
                'passes_attempted': stats.passes_attempted,
                'pass_accuracy_pct': self._calculate_percentage(stats.passes_completed, stats.passes_attempted),
                'distance_covered_km': float(stats.distance_covered),
                'sprints': stats.sprints,
                'top_speed_kmh': float(stats.top_speed),
                'tackles_won': stats.tackles_won,
                'tackles_attempted': stats.tackles_attempted,
                'tackle_success_pct': self._calculate_percentage(stats.tackles_won, stats.tackles_attempted),
                'interceptions': stats.interceptions,
                'clearances': stats.clearances,
                'shots_on_target': stats.shots_on_target,
                'shots_off_target': stats.shots_off_target,
                'shots_blocked': stats.shots_blocked,
                'total_shots': stats.shots_on_target + stats.shots_off_target + stats.shots_blocked,
                'shot_accuracy_pct': self._calculate_percentage(stats.shots_on_target, stats.shots_on_target + stats.shots_off_target),
                'crosses_completed': stats.crosses_completed,
                'crosses_attempted': stats.crosses_attempted,
                'cross_accuracy_pct': self._calculate_percentage(stats.crosses_completed, stats.crosses_attempted),
                'fouls_committed': stats.fouls_committed,
                'fouls_won': stats.fouls_won,
                'yellow_cards': stats.yellow_cards,
                'red_cards': stats.red_cards,
                'offsides': stats.offsides,
                'season': self._determine_season(stats.match.scheduled_datetime),
                'match_type': stats.match.match_type
            }
            performance_data.append(performance_record)
        
        return {
            'table_name': 'player_performance',
            'record_count': len(performance_data),
            'columns': list(performance_data[0].keys()) if performance_data else [],
            'data': performance_data
        }
    
    def _export_formations_data(self, date_range=None):
        lineups_query = MatchLineup.objects.select_related('formation', 'match', 'match__opponent')
        
        if date_range:
            start_date = datetime.fromisoformat(date_range['start_date'])
            end_date = datetime.fromisoformat(date_range['end_date'])
            lineups_query = lineups_query.filter(
                match__scheduled_datetime__date__range=[start_date.date(), end_date.date()]
            )
        
        formations_data = []
        for lineup in lineups_query:
            formation_record = {
                'lineup_id': str(lineup.id),
                'match_id': str(lineup.match.id),
                'match_date': lineup.match.scheduled_datetime.date().isoformat(),
                'formation_name': lineup.formation.name,
                'formation_id': str(lineup.formation.id),
                'opponent': lineup.match.opponent.name,
                'is_home': lineup.match.is_home,
                'match_result': lineup.match.result,
                'chelsea_score': lineup.match.chelsea_score,
                'opponent_score': lineup.match.opponent_score,
                'goal_difference': lineup.match.chelsea_score - lineup.match.opponent_score,
                'is_starting_eleven': lineup.is_starting_eleven,
                'formation_style': self._determine_formation_style(lineup.formation.name),
                'defensive_players': lineup.formation.defensive_line,
                'midfield_players': lineup.formation.midfield_line,
                'attacking_players': lineup.formation.attacking_line,
                'formation_effectiveness': self._calculate_formation_match_effectiveness(lineup),
                'season': self._determine_season(lineup.match.scheduled_datetime),
                'match_type': lineup.match.match_type
            }
            formations_data.append(formation_record)
        
        return {
            'table_name': 'formations',
            'record_count': len(formations_data),
            'columns': list(formations_data[0].keys()) if formations_data else [],
            'data': formations_data
        }
    
    def _export_analytics_data(self, date_range=None):
        analytics_query = Analytics.objects.select_related('created_by', 'related_match', 'related_player', 'related_formation')
        
        if date_range:
            start_date = datetime.fromisoformat(date_range['start_date'])
            end_date = datetime.fromisoformat(date_range['end_date'])
            analytics_query = analytics_query.filter(
                created_at__date__range=[start_date.date(), end_date.date()]
            )
        
        analytics_data = []
        for analytics in analytics_query:
            analytics_record = {
                'analytics_id': str(analytics.id),
                'analysis_type': analytics.analysis_type,
                'title': analytics.title,
                'description': analytics.description,
                'confidence_score': float(analytics.confidence_score),
                'created_date': analytics.created_at.date().isoformat(),
                'created_datetime': analytics.created_at.isoformat(),
                'created_by': analytics.created_by.username if analytics.created_by else 'System',
                'related_match_id': str(analytics.related_match.id) if analytics.related_match else None,
                'related_player_id': str(analytics.related_player.id) if analytics.related_player else None,
                'related_formation_id': str(analytics.related_formation.id) if analytics.related_formation else None,
                'insights_count': len(analytics.insights),
                'recommendations_count': len(analytics.recommendations),
                'data_points_available': bool(analytics.data_points),
                'season': self._determine_season(analytics.created_at) if analytics.created_at else None
            }
            analytics_data.append(analytics_record)
        
        return {
            'table_name': 'analytics',
            'record_count': len(analytics_data),
            'columns': list(analytics_data[0].keys()) if analytics_data else [],
            'data': analytics_data
        }
    
    def _export_team_stats_data(self, date_range=None):
        team_stats_query = TeamStats.objects.select_related('match', 'match__opponent')
        
        if date_range:
            start_date = datetime.fromisoformat(date_range['start_date'])
            end_date = datetime.fromisoformat(date_range['end_date'])
            team_stats_query = team_stats_query.filter(
                match__scheduled_datetime__date__range=[start_date.date(), end_date.date()]
            )
        
        team_stats_data = []
        for team_stats in team_stats_query:
            team_record = {
                'team_stats_id': str(team_stats.id),
                'match_id': str(team_stats.match.id),
                'match_date': team_stats.match.scheduled_datetime.date().isoformat(),
                'opponent': team_stats.match.opponent.name,
                'is_home': team_stats.match.is_home,
                'match_result': team_stats.match.result,
                'possession_percentage': float(team_stats.possession_percentage),
                'total_passes': team_stats.total_passes,
                'pass_accuracy': float(team_stats.pass_accuracy),
                'shots_total': team_stats.shots_total,
                'shots_on_target': team_stats.shots_on_target,
                'shots_off_target': team_stats.shots_off_target,
                'shot_accuracy_pct': self._calculate_percentage(team_stats.shots_on_target, team_stats.shots_total),
                'corners': team_stats.corners,
                'offsides': team_stats.offsides,
                'fouls_committed': team_stats.fouls_committed,
                'yellow_cards': team_stats.yellow_cards,
                'red_cards': team_stats.red_cards,
                'total_cards': team_stats.yellow_cards + team_stats.red_cards,
                'distance_covered_total_km': float(team_stats.distance_covered_total),
                'sprints_total': team_stats.sprints_total,
                'season': self._determine_season(team_stats.match.scheduled_datetime),
                'match_type': team_stats.match.match_type
            }
            team_stats_data.append(team_record)
        
        return {
            'table_name': 'team_statistics',
            'record_count': len(team_stats_data),
            'columns': list(team_stats_data[0].keys()) if team_stats_data else [],
            'data': team_stats_data
        }
    
    def export_to_csv(self, export_type='all', date_range=None):
        export_data = self.export_for_powerbi(export_type, date_range)
        csv_files = {}
        
        for data_source_name, data_source in export_data['data_sources'].items():
            if not data_source['data']:
                continue
            
            output = StringIO()
            writer = csv.DictWriter(output, fieldnames=data_source['columns'])
            writer.writeheader()
            writer.writerows(data_source['data'])
            
            csv_files[data_source_name] = output.getvalue()
            output.close()
        
        if len(csv_files) == 1:
            return list(csv_files.values())[0]
        
        return csv_files
    
    def export_to_excel(self, export_type='all', date_range=None):
        export_data = self.export_for_powerbi(export_type, date_range)
        
        workbook = Workbook()
        workbook.remove(workbook.active)
        
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
        
        for data_source_name, data_source in export_data['data_sources'].items():
            if not data_source['data']:
                continue
            
            worksheet = workbook.create_sheet(title=data_source_name[:31])
            
            if data_source['data']:
                df = pd.DataFrame(data_source['data'])
                
                for row in dataframe_to_rows(df, index=False, header=True):
                    worksheet.append(row)
                
                for cell in worksheet[1]:
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = Alignment(horizontal="center")
                
                for column in worksheet.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    
                    adjusted_width = min(max_length + 2, 50)
                    worksheet.column_dimensions[column_letter].width = adjusted_width
        
        output = BytesIO()
        workbook.save(output)
        return output.getvalue()
    
    def schedule_export(self, schedule_config):
        try:
            export_schedule = {
                'schedule_id': f"export_{timezone.now().strftime('%Y%m%d_%H%M%S')}",
                'export_type': schedule_config.get('export_type', 'all'),
                'frequency': schedule_config.get('frequency', 'weekly'),
                'time_of_day': schedule_config.get('time_of_day', '06:00'),
                'format': schedule_config.get('format', 'powerbi'),
                'recipients': schedule_config.get('recipients', []),
                'created_at': timezone.now().isoformat(),
                'next_execution': self._calculate_next_execution(
                    schedule_config.get('frequency', 'weekly'),
                    schedule_config.get('time_of_day', '06:00')
                ).isoformat()
            }
            
            self._save_export_schedule(export_schedule)
            
            return {
                'success': True,
                'schedule_id': export_schedule['schedule_id'],
                'next_execution': export_schedule['next_execution']
            }
            
        except Exception as e:
            self.logger.error(f"Failed to schedule export: {str(e)}")
            return {'success': False, 'error': str(e)}
    
    def _categorise_position(self, position):
        categories = {
            'GK': 'Goalkeeper',
            'CB': 'Defender',
            'LB': 'Defender',
            'RB': 'Defender',
            'CDM': 'Midfielder',
            'CM': 'Midfielder',
            'CAM': 'Midfielder',
            'LM': 'Midfielder',
            'RM': 'Midfielder',
            'LW': 'Forward',
            'RW': 'Forward',
            'ST': 'Forward'
        }
        return categories.get(position, 'Unknown')
    
    def _calculate_percentage(self, numerator, denominator):
        if denominator == 0:
            return 0
        return round((numerator / denominator) * 100, 2)
    
    def _determine_season(self, match_datetime):
        year = match_datetime.year
        if match_datetime.month >= 7:
            return f"{year}/{year + 1}"
        else:
            return f"{year - 1}/{year}"
    
    def _determine_formation_style(self, formation_name):
        attacking_formations = ['4-3-3', '3-4-3', '4-2-3-1']
        defensive_formations = ['5-3-2', '5-4-1']
        counter_attacking = ['3-5-2']
        
        if formation_name in attacking_formations:
            return 'Attacking'
        elif formation_name in defensive_formations:
            return 'Defensive'
        elif formation_name in counter_attacking:
            return 'Counter-Attacking'
        else:
            return 'Balanced'
    
    def _calculate_formation_match_effectiveness(self, lineup):
        match = lineup.match
        
        if match.result == 'WIN':
            base_score = 100
        elif match.result == 'DRAW':
            base_score = 60
        else:
            base_score = 20
        
        goal_difference = match.chelsea_score - match.opponent_score
        goal_modifier = goal_difference * 10
        
        effectiveness = max(0, min(100, base_score + goal_modifier))
        return round(effectiveness, 2)
    
    def _save_powerbi_export(self, export_data, export_type):
        try:
            filename = f"powerbi_export_{export_type}_{timezone.now().strftime('%Y%m%d_%H%M%S')}.json"
            filepath = self.export_path / 'powerbi' / 'datasets' / filename
            
            filepath.parent.mkdir(parents=True, exist_ok=True)
            
            with open(filepath, 'w', encoding='utf-8') as f:
                json.dump(export_data, f, indent=2, ensure_ascii=False)
            
            self.logger.info(f"PowerBI export saved: {filepath}")
            
        except Exception as e:
            self.logger.error(f"Failed to save PowerBI export: {str(e)}")
    
    def _calculate_next_execution(self, frequency, time_of_day):
        now = timezone.now()
        hour, minute = map(int, time_of_day.split(':'))
        
        next_execution = now.replace(hour=hour, minute=minute, second=0, microsecond=0)
        
        if frequency == 'daily':
            if next_execution <= now:
                next_execution += timedelta(days=1)
        elif frequency == 'weekly':
            days_until_sunday = (6 - now.weekday()) % 7
            if days_until_sunday == 0 and next_execution <= now:
                days_until_sunday = 7
            next_execution += timedelta(days=days_until_sunday)
        elif frequency == 'monthly':
            if next_execution.day == 1:
                if next_execution <= now:
                    if next_execution.month == 12:
                        next_execution = next_execution.replace(year=next_execution.year + 1, month=1)
                    else:
                        next_execution = next_execution.replace(month=next_execution.month + 1)
            else:
                next_execution = next_execution.replace(day=1)
                if next_execution <= now:
                    if next_execution.month == 12:
                        next_execution = next_execution.replace(year=next_execution.year + 1, month=1)
                    else:
                        next_execution = next_execution.replace(month=next_execution.month + 1)
        
        return next_execution
    
    def _save_export_schedule(self, export_schedule):
        try:
            filename = f"export_schedule_{export_schedule['schedule_id']}.json"
            filepath = self.export_path / 'schedules' / filename
            
            filepath.parent.mkdir(parents=True, exist_ok=True)
            
            with open(filepath, 'w', encoding='utf-8') as f:
                json.dump(export_schedule, f, indent=2, ensure_ascii=False)
            
            self.logger.info(f"Export schedule saved: {filepath}")
            
        except Exception as e:
            self.logger.error(f"Failed to save export schedule: {str(e)}")
    
    def get_export_summary(self, days=30):
        cutoff_date = timezone.now() - timedelta(days=days)
        
        summary = {
            'period_days': days,
            'players_active': Player.objects.filter(is_active=True).count(),
            'matches_analyzed': Match.objects.filter(
                scheduled_datetime__gte=cutoff_date,
                status__in=['COMPLETED', 'FULL_TIME']
            ).count(),
            'performance_records': PlayerStats.objects.filter(
                match__scheduled_datetime__gte=cutoff_date
            ).count(),
            'formations_used': MatchLineup.objects.filter(
                match__scheduled_datetime__gte=cutoff_date
            ).values('formation__name').distinct().count(),
            'analytics_generated': Analytics.objects.filter(
                created_at__gte=cutoff_date
            ).count(),
            'data_quality_indicators': self._assess_data_quality(cutoff_date)
        }
        
        return summary
    
    def _assess_data_quality(self, cutoff_date):
        completed_matches = Match.objects.filter(
            scheduled_datetime__gte=cutoff_date,
            status__in=['COMPLETED', 'FULL_TIME']
        )
        
        matches_with_stats = completed_matches.filter(
            player_stats__isnull=False
        ).distinct().count()
        
        matches_with_team_stats = completed_matches.filter(
            team_stats__isnull=False
        ).count()
        
        matches_with_lineups = completed_matches.filter(
            lineups__isnull=False
        ).distinct().count()
        
        total_matches = completed_matches.count()
        
        return {
            'matches_with_player_stats_pct': self._calculate_percentage(matches_with_stats, total_matches),
            'matches_with_team_stats_pct': self._calculate_percentage(matches_with_team_stats, total_matches),
            'matches_with_lineups_pct': self._calculate_percentage(matches_with_lineups, total_matches),
            'overall_data_completeness_pct': self._calculate_percentage(
                min(matches_with_stats, matches_with_team_stats, matches_with_lineups),
                total_matches
            )
        }